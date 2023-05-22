import pyautogui
from pyautogui import *
import pyperclip
import openpyxl
from openpyxl.styles import PatternFill

# https://chrome.google.com/webstore/detail/recaptcha-autoclick/caahalkghnhbabknipmconmbicpkcopl
# https://chrome.google.com/webstore/detail/buster-captcha-solver-for/mpbjkejclgfgadiemmefgebjfooflfhl

wb = openpyxl.load_workbook('dados_2024_5.xlsx', data_only=True, )

'''
sleep(5)
print(pyautogui.position())
input()
'''

# get first worksheet
ws = wb.worksheets[0]
numeros_precarios = []

# check first column in first 10 rows for fill color
for row in range(2, ws.max_row + 1):
    codigo_item = ws.cell(column=1, row=row)
    texto_longo = ws.cell(column=2, row=row)
    bgColor = codigo_item.fill.bgColor.index
    fgColor = codigo_item.fill.fgColor.index
    if bgColor == '00000000' or fgColor == '00000000':
        if codigo_item.value is not None:
            numeros_precarios.append(codigo_item.value)
            # numeros_precarios.append(texto_longo.value)

print(numeros_precarios)

sleep(0.5)


pyautogui.click(x=720, y=1067)
sleep(1)

for indice, numero in enumerate(numeros_precarios):
    print(numero)
    pyperclip.copy('a')
    contador = 0
    while str(pyperclip.paste()) != 'https://web.trf3.jus.br/consultas/Internet/ConsultaReqPag/Consultar':
        if contador == 10:
            exit()
        pyautogui.click(x=538, y=73)
        pyautogui.sleep(0.5)
        pyautogui.press('backspace')
        sleep(1)
        pyautogui.write('https://web.trf3.jus.br/consultas/Internet/ConsultaReqPag')
        sleep(0.5)
        pyautogui.press('enter')
        sleep(1)
        pyautogui.click(x=690, y=683)
        sleep(0.5)
        pyautogui.write(f'{str(numero)}')
        sleep(1)
        pyautogui.click(x=383, y=734)
        sleep(1)
        # pyautogui.click(x=592, y=994)
        pyautogui.click(x=611, y=940)
        sleep(10)
        pyautogui.click(x=676, y=868)
        sleep(1)
        pyautogui.click(x=501, y=75)
        sleep(0.5)
        pyautogui.keyDown('ctrl')
        sleep(0.5)
        pyautogui.press('c')
        sleep(0.5)
        pyautogui.keyUp('ctrl')
        sleep(0.5)
        print(pyperclip.paste())
        pyautogui.press('backspace')
        contador += 1
    sleep(1)
    pyautogui.click(x=1890, y=83)
    sleep(1)
    pyautogui.click(x=1598, y=394)
    sleep(1)
    pyautogui.click(x=1515, y=932)
    sleep(1.5)
    pyautogui.click(x=971, y=124)
    sleep(1)
    pyautogui.write(r'C:\Users\maoliveira\OneDrive - Trident Energy Management LTD\Documents\My files\free\Marcio\dados_2024')
    sleep(1)
    pyautogui.press('enter')
    sleep(1)
    pyautogui.click(x=848, y=606)
    sleep(1)
    pyautogui.write(f'{str(numero)}')
    sleep(1)
    pyautogui.press('enter')
    sleep(0.5)
    pyautogui.click(x=1068, y=548)
    sleep(0.5)