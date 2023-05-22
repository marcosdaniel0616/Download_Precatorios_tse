import pyautogui
from pyautogui import *
import pyperclip
import openpyxl
from openpyxl.styles import PatternFill

# https://chrome.google.com/webstore/detail/recaptcha-autoclick/caahalkghnhbabknipmconmbicpkcopl
# https://chrome.google.com/webstore/detail/buster-captcha-solver-for/mpbjkejclgfgadiemmefgebjfooflfhl

wb = openpyxl.load_workbook('dados_2024_4.xlsx', data_only=True, )

'''
sleep(5)
print(pyautogui.position())
input()
'''

# get first worksheet
ws = wb.worksheets[0]
numeros_precarios = []

# check first column in first 10 rows for fill color
for row in range(2, ws.max_row + 1):
    codigo_item = ws.cell(column=1, row=row)
    texto_longo = ws.cell(column=2, row=row)
    bgColor = codigo_item.fill.bgColor.index
    fgColor = codigo_item.fill.fgColor.index
    if bgColor == '00000000' or fgColor == '00000000':
        if codigo_item.value is not None:
            numeros_precarios.append(codigo_item.value)
            # numeros_precarios.append(texto_longo.value)

print(numeros_precarios)

sleep(0.5)


pyautogui.click(x=720, y=1067)
sleep(1)

for indice, numero in enumerate(numeros_precarios):
    print(numero)
    pyperclip.copy('a')
    contador = 0
    while str(pyperclip.paste()) != 'https://web.trf3.jus.br/consultas/Internet/ConsultaReqPag/Consultar':
        if contador == 10:
            exit()
        pyautogui.click(x=538, y=73)
        pyautogui.sleep(1)
        pyautogui.press('backspace')
        sleep(1)
        pyautogui.write('https://web.trf3.jus.br/consultas/Internet/ConsultaReqPag')
        sleep(1)
        pyautogui.press('enter')
        sleep(1)
        pyautogui.click(x=727, y=680)
        sleep(1)
        pyautogui.write(f'{str(numero)}')
        sleep(1)
        pyautogui.click(x=443, y=758)
        sleep(1)
        pyautogui.click(x=600, y=933)
        sleep(10)
        pyautogui.click(x=665, y=865)
        sleep(1)
        pyautogui.click(x=538, y=73)
        sleep(0.5)
        pyautogui.keyDown('ctrl')
        sleep(0.5)
        pyautogui.press('c')
        sleep(0.5)
        pyautogui.keyUp('ctrl')
        sleep(0.5)
        print(pyperclip.paste())
        pyautogui.press('backspace')
        contador += 1
    sleep(1)
    pyautogui.click(x=1810, y=74)
    sleep(1)
    pyautogui.click(x=1412, y=719)
    sleep(1)
    pyautogui.click(x=217, y=845)
    sleep(1.5)
    pyautogui.click(x=946, y=214)
    sleep(1)
    pyautogui.write(r'C:\Users\maoliveira\OneDrive - Trident Energy Management LTD\Documents\My files\free\Marcio\dados_2024')
    sleep(1)
    pyautogui.press('enter')
    sleep(1)
    pyautogui.click(x=838, y=700)
    sleep(1)
    pyautogui.write(f'{str(numero)}')
    sleep(1)
    pyautogui.press('enter')
    sleep(0.5)
    pyautogui.click(x=1034, y=547)
    sleep(0.5)


